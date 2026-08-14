#!/usr/bin/env python3
"""Detect duplicate names, headers, rows, and configured keys in an OOXML workbook."""

from __future__ import annotations

import argparse
import json
import math
import numbers
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


SHEET_ROLES = {"formula_output", "lookup", "other", "pivot", "raw", "supplemental"}


def load_json(path: Path) -> dict:
    with path.open(encoding="utf-8") as stream:
        value = json.load(stream)
    if not isinstance(value, dict):
        raise ValueError(f"Expected a JSON object in {path}")
    return value


def normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", text.strip()).casefold()


def valid_timestamp(value: object) -> bool:
    if not isinstance(value, str) or not value.strip():
        return False
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return False
    return parsed.tzinfo is not None


def is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def exact_value_key(value: object) -> tuple[str, object]:
    if is_blank(value):
        return "blank", ""
    if isinstance(value, bool):
        return "bool", value
    if isinstance(value, numbers.Number):
        numeric = float(value)
        if math.isnan(numeric):
            return "blank", ""
        return "number", numeric
    if isinstance(value, (datetime, date)):
        return "date", value.isoformat()
    if isinstance(value, str):
        return "text", value
    return type(value).__name__, str(value)


def business_key_value(value: object, normalize_key_text: bool) -> tuple[str, object]:
    kind, normalized = exact_value_key(value)
    if kind == "text" and normalize_key_text:
        normalized = normalize_text(normalized)
    return kind, normalized


def json_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def trimmed_row(values: tuple[object, ...]) -> list[object]:
    result = list(values)
    while result and is_blank(result[-1]):
        result.pop()
    return result


def column_positions(headers: list[object]) -> dict[str, list[int]]:
    result: dict[str, list[int]] = defaultdict(list)
    for position, header in enumerate(headers):
        if not is_blank(header):
            result[normalize_text(header)].append(position)
    return result


def resolve_columns(
    requested: list[str], positions: dict[str, list[int]], label: str
) -> tuple[list[int], list[str]]:
    resolved = []
    errors = []
    for column in requested:
        hits = positions.get(normalize_text(column), [])
        if not hits:
            errors.append(f"{label} column not found: {column}")
        elif len(hits) > 1:
            errors.append(f"{label} column is ambiguous: {column}")
        else:
            resolved.append(hits[0])
    return resolved, errors


def add_seen_group(
    seen: dict[tuple, int], groups: dict[tuple, list[int]], key: tuple, row: int
) -> None:
    if key in groups:
        groups[key].append(row)
    elif key in seen:
        groups[key] = [seen[key], row]
    else:
        seen[key] = row


def inspect_sheet(worksheet, config: dict, default_sample_limit: int) -> dict:
    name = worksheet.title
    header_row = int(config.get("header_row", 1))
    data_start_row = int(config.get("data_start_row", header_row + 1))
    max_data_row = config.get("max_data_row")
    sample_limit = int(config.get("sample_limit", default_sample_limit))
    if header_row < 1 or data_start_row <= header_row:
        raise ValueError(
            f"{name}: header_row must be positive and data_start_row must follow it"
        )

    header_values = next(
        worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
        (),
    )
    headers = trimmed_row(header_values)
    positions = column_positions(headers)
    duplicate_headers = [
        {
            "normalized_header": normalized,
            "columns": [position + 1 for position in hits],
            "labels": [json_value(headers[position]) for position in hits],
        }
        for normalized, hits in positions.items()
        if len(hits) > 1
    ]

    key_columns = config.get("key_columns", [])
    ignore_columns = config.get("ignore_columns", [])
    if not isinstance(key_columns, list) or not all(
        isinstance(item, str) for item in key_columns
    ):
        raise ValueError(f"{name}: key_columns must be a list of strings")
    if not isinstance(ignore_columns, list) or not all(
        isinstance(item, str) for item in ignore_columns
    ):
        raise ValueError(f"{name}: ignore_columns must be a list of strings")

    key_positions, key_errors = resolve_columns(key_columns, positions, "key")
    ignored_positions, ignore_errors = resolve_columns(
        ignore_columns, positions, "ignored"
    )
    resolution_errors = key_errors + ignore_errors
    exact_positions = [
        position
        for position in range(len(headers))
        if position not in set(ignored_positions)
    ]

    normalize_key_text = bool(config.get("normalize_key_text", False))
    require_complete_key = bool(config.get("require_complete_key", True))
    allowed_raw = config.get("allowed_duplicate_keys", [])
    if not isinstance(allowed_raw, list):
        raise ValueError(f"{name}: allowed_duplicate_keys must be a list")
    allowed_keys = set()
    allowed_key_evidence = []
    for item in allowed_raw:
        if not isinstance(item, dict):
            raise ValueError(
                f"{name}: each allowed duplicate key must be an evidence object"
            )
        values = item.get("values")
        if not isinstance(values, list) or len(values) != len(key_columns):
            raise ValueError(
                f"{name}: each allowed duplicate key must contain {len(key_columns)} values"
            )
        for field in ("reason", "owner", "evidence"):
            if not isinstance(item.get(field), str) or not item[field].strip():
                raise ValueError(f"{name}: allowed duplicate key {field} is required")
        if not valid_timestamp(item.get("checked_at")):
            raise ValueError(
                f"{name}: allowed duplicate key checked_at must be ISO-8601"
            )
        allowed_keys.add(
            tuple(business_key_value(value, normalize_key_text) for value in values)
        )
        allowed_key_evidence.append(item)

    exact_seen: dict[tuple, int] = {}
    exact_groups: dict[tuple, list[int]] = {}
    key_seen: dict[tuple, int] = {}
    key_groups: dict[tuple, list[int]] = {}
    key_display: dict[tuple, list[object]] = {}
    incomplete_key_rows = []
    incomplete_key_row_count = 0
    populated_rows = 0
    check_exact = bool(config.get("check_exact_rows", True))

    max_col = max(1, len(headers))
    for row_number, values in enumerate(
        worksheet.iter_rows(
            min_row=data_start_row,
            max_row=max_data_row,
            max_col=max_col,
            values_only=True,
        ),
        start=data_start_row,
    ):
        values = tuple(values[: len(headers)])
        if all(is_blank(value) for value in values):
            continue
        populated_rows += 1

        if check_exact and exact_positions:
            exact_key = tuple(exact_value_key(values[pos]) for pos in exact_positions)
            add_seen_group(exact_seen, exact_groups, exact_key, row_number)

        if key_positions and not resolution_errors:
            raw_key = [values[position] for position in key_positions]
            if any(is_blank(value) for value in raw_key):
                incomplete_key_row_count += 1
                if require_complete_key and len(incomplete_key_rows) < sample_limit:
                    incomplete_key_rows.append(row_number)
                continue
            key = tuple(
                business_key_value(value, normalize_key_text) for value in raw_key
            )
            key_display.setdefault(key, [json_value(value) for value in raw_key])
            add_seen_group(key_seen, key_groups, key, row_number)

    duplicate_exact_groups = [
        {"rows": rows}
        for rows in sorted(exact_groups.values(), key=lambda item: item[0])[
            :sample_limit
        ]
    ]
    duplicate_key_groups = []
    allowed_duplicate_key_groups = []
    for key, rows in sorted(key_groups.items(), key=lambda item: item[1][0]):
        entry = {
            "key": {
                column: value
                for column, value in zip(key_columns, key_display[key], strict=True)
            },
            "rows": rows,
        }
        target = (
            allowed_duplicate_key_groups
            if key in allowed_keys
            else duplicate_key_groups
        )
        if len(target) < sample_limit:
            target.append(entry)

    checks_duplicate_headers = bool(config.get("check_duplicate_headers", True))
    blocking_duplicate_headers = duplicate_headers if checks_duplicate_headers else []
    issues = list(resolution_errors)
    if blocking_duplicate_headers:
        issues.append(f"{len(blocking_duplicate_headers)} duplicate header group(s)")
    if exact_groups:
        issues.append(f"{len(exact_groups)} exact duplicate row group(s)")
    unexpected_key_group_count = sum(key not in allowed_keys for key in key_groups)
    if unexpected_key_group_count:
        issues.append(
            f"{unexpected_key_group_count} unexpected duplicate business-key group(s)"
        )
    if require_complete_key and incomplete_key_row_count:
        issues.append(
            f"{incomplete_key_row_count} populated row(s) have incomplete business keys"
        )

    return {
        "name": name,
        "header_row": header_row,
        "data_start_row": data_start_row,
        "headers": [json_value(value) for value in headers],
        "populated_data_rows": populated_rows,
        "duplicate_header_groups": duplicate_headers,
        "duplicate_header_check_enabled": checks_duplicate_headers,
        "exact_row_check_enabled": check_exact,
        "exact_duplicate_group_count": len(exact_groups),
        "exact_duplicate_groups_sample": duplicate_exact_groups,
        "key_columns": key_columns,
        "duplicate_key_group_count": len(key_groups),
        "unexpected_duplicate_key_group_count": unexpected_key_group_count,
        "duplicate_key_groups_sample": duplicate_key_groups,
        "allowed_duplicate_key_groups_sample": allowed_duplicate_key_groups,
        "allowed_duplicate_key_evidence": allowed_key_evidence,
        "incomplete_key_row_count": incomplete_key_row_count,
        "incomplete_key_rows_sample": incomplete_key_rows,
        "issues": issues,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--config", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    args = parser.parse_args()

    if args.workbook.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise SystemExit("Duplicate checking currently supports OOXML workbooks only")
    config = load_json(args.config)
    sample_limit = int(config.get("sample_limit", 25))
    sheet_configs = config.get("sheets", [])
    if not isinstance(sheet_configs, list) or not sheet_configs:
        raise SystemExit("Config must contain a non-empty sheets list")

    workbook = load_workbook(args.workbook, read_only=True, data_only=False)
    try:
        normalized_names: dict[str, list[str]] = defaultdict(list)
        for name in workbook.sheetnames:
            normalized_names[normalize_text(name)].append(name)
        sheet_name_collisions = [
            {"normalized_name": normalized, "sheets": names}
            for normalized, names in normalized_names.items()
            if len(names) > 1
        ]
        configured_names = []
        sheet_reports = []
        top_level_issues = []
        if config.get("check_normalized_sheet_names") is not True:
            top_level_issues.append(
                "check_normalized_sheet_names must be true for a release audit"
            )
        for sheet_config in sheet_configs:
            if not isinstance(sheet_config, dict) or not isinstance(
                sheet_config.get("name"), str
            ):
                raise SystemExit("Every sheets entry must be an object with a name")
            name = sheet_config["name"]
            configured_names.append(name)
            role = sheet_config.get("role")
            if role not in SHEET_ROLES:
                top_level_issues.append(
                    f"{name}: role must be one of {sorted(SHEET_ROLES)}"
                )
            for field in ("check_duplicate_headers", "check_exact_rows"):
                if sheet_config.get(field) not in {True, False}:
                    top_level_issues.append(f"{name}: {field} must be boolean")
            if role in {
                "formula_output",
                "lookup",
                "raw",
                "supplemental",
            } and not sheet_config.get("key_columns"):
                top_level_issues.append(
                    f"{name}: role {role} requires non-empty key_columns"
                )
            if name not in workbook.sheetnames:
                top_level_issues.append(f"configured sheet not found: {name}")
                continue
            sheet_reports.append(
                inspect_sheet(workbook[name], sheet_config, sample_limit)
            )
        repeated_configs = sorted(
            name for name in set(configured_names) if configured_names.count(name) > 1
        )
        if repeated_configs:
            top_level_issues.append(
                f"sheet configured more than once: {', '.join(repeated_configs)}"
            )
        excluded = config.get("excluded_sheets", [])
        if not isinstance(excluded, list):
            raise SystemExit("excluded_sheets must be a list")
        excluded_names = []
        for index, item in enumerate(excluded):
            if not isinstance(item, dict) or not isinstance(item.get("name"), str):
                top_level_issues.append(
                    f"excluded_sheets[{index}] must be an object with a name"
                )
                continue
            excluded_names.append(item["name"])
            if item.get("role") not in {"other", "pivot"}:
                top_level_issues.append(
                    f"excluded_sheets[{index}].role must be other or pivot"
                )
            for field in ("reason", "owner", "evidence"):
                if not isinstance(item.get(field), str) or not item[field].strip():
                    top_level_issues.append(
                        f"excluded_sheets[{index}].{field} is required"
                    )
            if not valid_timestamp(item.get("checked_at")):
                top_level_issues.append(
                    f"excluded_sheets[{index}].checked_at must be ISO-8601"
                )
        repeated_exclusions = sorted(
            name for name in set(excluded_names) if excluded_names.count(name) > 1
        )
        if repeated_exclusions:
            top_level_issues.append(
                f"sheet excluded more than once: {', '.join(repeated_exclusions)}"
            )
        overlap = sorted(set(configured_names) & set(excluded_names))
        if overlap:
            top_level_issues.append(
                f"sheet cannot be both configured and excluded: {', '.join(overlap)}"
            )
        unknown_exclusions = sorted(set(excluded_names) - set(workbook.sheetnames))
        if unknown_exclusions:
            top_level_issues.append(
                f"excluded sheet not found: {', '.join(unknown_exclusions)}"
            )
        uncovered_sheets = sorted(
            set(workbook.sheetnames) - set(configured_names) - set(excluded_names)
        )
        if uncovered_sheets:
            top_level_issues.append(
                f"workbook sheets missing from the audit contract: {', '.join(uncovered_sheets)}"
            )
        if sheet_name_collisions:
            top_level_issues.append(
                f"{len(sheet_name_collisions)} normalized sheet-name collision(s)"
            )
        blockers = top_level_issues + [
            f"{sheet['name']}: {issue}"
            for sheet in sheet_reports
            for issue in sheet["issues"]
        ]
        report = {
            "check": "workbook_duplicates",
            "passed": not blockers,
            "workbook": str(args.workbook.resolve()),
            "config": str(args.config.resolve()),
            "sheet_count": len(workbook.sheetnames),
            "configured_sheet_count": len(sheet_configs),
            "excluded_sheets": excluded,
            "uncovered_sheets": uncovered_sheets,
            "normalized_sheet_name_check_enabled": config.get(
                "check_normalized_sheet_names"
            )
            is True,
            "normalized_sheet_name_collisions": sheet_name_collisions,
            "sheets": sheet_reports,
            "blockers": blockers,
        }
    finally:
        workbook.close()

    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(
        json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "passed": report["passed"],
                "sheets_checked": len(report["sheets"]),
                "blockers": len(report["blockers"]),
            }
        )
    )
    return 0 if report["passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
