#!/usr/bin/env python3
"""Run adversarial regression checks for the MEGA onboarding skill suite."""

from __future__ import annotations

import csv
import hashlib
import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook


SUITE_ROOT = Path(__file__).resolve().parents[2]
BOOST = SUITE_ROOT / "mega-boost-onboarding" / "scripts"
OVERCOUNTING = SUITE_ROOT / "mega-boost-overcounting" / "scripts"
SUBNATIONAL = SUITE_ROOT / "mega-subnational-onboarding" / "scripts"
VALIDATION = SUITE_ROOT / "mega-onboarding-validation" / "scripts"
NOW = "2026-01-01T00:00:00Z"


def write_json(path: Path, value: object) -> None:
    path.write_text(
        json.dumps(value, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )


def write_csv(path: Path, rows: list[list[object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as stream:
        csv.writer(stream).writerows(rows)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def execute(arguments: list[object]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, "-B", *map(str, arguments)],
        text=True,
        capture_output=True,
        check=False,
    )


def record(
    results: list[dict],
    name: str,
    completed: subprocess.CompletedProcess[str],
    expected: int,
) -> None:
    results.append(
        {
            "name": name,
            "passed": (completed.returncode == 0) == (expected == 0),
            "expected_exit_class": "zero" if expected == 0 else "nonzero",
            "actual_exit_code": completed.returncode,
            "stdout": completed.stdout.strip(),
            "stderr": completed.stderr.strip(),
        }
    )


def record_condition(
    results: list[dict],
    name: str,
    passed: bool,
    completed: subprocess.CompletedProcess[str],
) -> None:
    results.append(
        {
            "name": name,
            "passed": passed,
            "expected_exit_class": "condition",
            "actual_exit_code": completed.returncode,
            "stdout": completed.stdout.strip(),
            "stderr": completed.stderr.strip(),
        }
    )


def manifest_tests(root: Path, results: list[dict]) -> None:
    source = root / "source.xlsx"
    source.write_bytes(b"synthetic-source")
    source_inventory = root / "source-inventory.json"
    write_json(
        source_inventory,
        {
            "country": {"name": "Auditland", "iso2": "AU", "iso3": "AUD"},
            "captured_at": NOW,
            "scope": {
                "stages": ["approved", "executed"],
                "expected_years": [2023],
                "currency": "LCU",
                "amount_unit": "units",
                "fiscal_year_convention": "Calendar year",
                "subnational_review": "central_only",
                "published_products": ["boost_country_gold"],
            },
            "primary_source_ids": ["boost_workbook"],
            "sources": [
                {
                    "id": "boost_workbook",
                    "kind": "boost_workbook",
                    "authority": "authoritative",
                    "required": True,
                    "path_or_url": str(source),
                    "local_snapshot": str(source),
                    "sha256": sha256(source),
                    "owner": "regression-test",
                    "format": "xlsx",
                    "data_classification": "internal",
                    "derived_from": [],
                }
            ],
        },
    )
    for repo in ("mega-boost", "mega-indicators", "rpf-country-dash"):
        (root / repo).mkdir()
    checker = SUITE_ROOT / "mega-country-onboarding" / "scripts" / "check_manifest.py"
    gate_names = {
        "intake",
        "workbook_audit",
        "workbook_duplicates",
        "overcounting",
        "foreign_funding",
        "boost_etl",
        "subnational",
        "cross_country",
        "dashboard",
        "staging",
        "production",
    }
    required_report_checks = {
        "intake": "source_intake",
        "workbook_duplicates": "workbook_duplicates",
        "overcounting": "overcounting",
        "foreign_funding": "foreign_funding",
        "subnational": "subnational_decision",
        "cross_country": "reconciliation",
    }
    evidence_by_gate = {}
    for gate in gate_names:
        evidence_file = root / f"report-{gate}.json"
        report = {
            "check": required_report_checks.get(gate, "generic"),
            "passed": True,
        }
        if gate == "intake":
            report["inventory_sha256"] = sha256(source_inventory)
        write_json(
            evidence_file,
            report,
        )
        evidence_by_gate[gate] = [
            {
                "kind": "file",
                "path": str(evidence_file),
                "sha256": sha256(evidence_file),
                "checked_at": NOW,
            }
        ]
    base = {
        "schema_version": 4,
        "country": {"name": "Auditland", "iso2": "AU", "iso3": "AUD"},
        "workspace": {"root": str(root), "manifest_path": "manifest.json"},
        "source_workbook": {
            "path_or_url": str(source),
            "local_snapshot": str(source),
            "sha256": sha256(source),
        },
        "source_package": {
            "inventory_path": str(source_inventory),
            "inventory_sha256": sha256(source_inventory),
            "primary_source_ids": ["boost_workbook"],
        },
        "repositories": {
            repo: {
                "path": str(root / repo),
                "baseline_ref": "baseline-sha",
                "final_ref": "final-sha",
            }
            for repo in ("mega-boost", "mega-indicators", "rpf-country-dash")
        },
        "subnational": {
            "required": False,
            "decision_evidence": ["central-only source review"],
        },
        "risks": [],
        "handoff": {"summary": "Synthetic regression fixture."},
    }
    ready = dict(base)
    ready["gates"] = {
        name: {"status": "passed", "evidence": evidence_by_gate[name]}
        for name in gate_names
    }
    ready_path = root / "manifest.json"
    write_json(ready_path, ready)
    record(
        results,
        "manifest accepts verifiable passed evidence",
        execute([checker, "check", "--manifest", ready_path, "--ready"]),
        0,
    )
    bad_hash = json.loads(json.dumps(ready))
    bad_hash["gates"]["workbook_audit"]["evidence"][0]["sha256"] = "0" * 64
    bad_hash_path = root / "manifest-bad-hash.json"
    bad_hash["workspace"]["manifest_path"] = str(bad_hash_path)
    write_json(bad_hash_path, bad_hash)
    record(
        results,
        "manifest rejects mismatched evidence hashes",
        execute([checker, "check", "--manifest", bad_hash_path, "--ready"]),
        1,
    )
    stale_intake = json.loads(json.dumps(ready))
    stale_intake_path = root / "manifest-stale-intake.json"
    stale_intake_report = root / "report-stale-intake.json"
    write_json(
        stale_intake_report,
        {
            "check": "source_intake",
            "passed": True,
            "inventory_sha256": "f" * 64,
        },
    )
    stale_intake["workspace"]["manifest_path"] = str(stale_intake_path)
    stale_intake["gates"]["intake"]["evidence"] = [
        {
            "kind": "file",
            "path": str(stale_intake_report),
            "sha256": sha256(stale_intake_report),
            "checked_at": NOW,
        }
    ]
    write_json(stale_intake_path, stale_intake)
    record(
        results,
        "manifest rejects an intake report from a different inventory",
        execute([checker, "check", "--manifest", stale_intake_path, "--ready"]),
        1,
    )
    bypass = dict(base)
    bypass["gates"] = {
        name: {
            "status": "not_applicable",
            "evidence": [
                {
                    "kind": "decision",
                    "owner": "tester",
                    "reason": "attempted bypass",
                    "checked_at": NOW,
                }
            ],
        }
        for name in gate_names
    }
    bypass_path = root / "manifest-bypass.json"
    write_json(bypass_path, bypass)
    record(
        results,
        "manifest rejects required not-applicable gates",
        execute([checker, "check", "--manifest", bypass_path, "--ready"]),
        1,
    )
    early_subnational = json.loads(json.dumps(ready))
    early_subnational_path = root / "manifest-early-subnational.json"
    early_subnational["workspace"]["manifest_path"] = str(early_subnational_path)
    early_subnational["gates"]["subnational"] = {
        "status": "not_started",
        "evidence": [],
        "next_action": "Decide central-only versus subnational.",
    }
    early_subnational["gates"]["boost_etl"] = {
        "status": "not_started",
        "evidence": [],
        "next_action": "Build BOOST ETL.",
    }
    write_json(early_subnational_path, early_subnational)
    next_run = execute([checker, "next", "--manifest", early_subnational_path])
    next_result = json.loads(next_run.stdout)
    record_condition(
        results,
        "manifest routes subnational work before BOOST ETL",
        next_run.returncode == 0 and next_result.get("current_gate") == "subnational",
        next_run,
    )


def source_inventory_tests(root: Path, results: list[dict]) -> None:
    checker = (
        SUITE_ROOT / "mega-country-onboarding" / "scripts" / "check_source_inventory.py"
    )
    workbook = root / "intake-workbook.xlsx"
    workbook.write_bytes(b"source-inventory-fixture")
    inventory = {
        "country": {"name": "Auditland", "iso2": "AU", "iso3": "AUD"},
        "captured_at": NOW,
        "scope": {
            "stages": ["approved", "executed"],
            "expected_years": [2023],
            "currency": "LCU",
            "amount_unit": "units",
            "fiscal_year_convention": "Calendar year",
            "subnational_review": "pending",
            "published_products": ["boost_country_gold", "country_dashboard"],
        },
        "primary_source_ids": ["boost_workbook"],
        "sources": [
            {
                "id": "boost_workbook",
                "kind": "boost_workbook",
                "authority": "authoritative",
                "required": True,
                "path_or_url": "https://example.test/source.xlsx",
                "local_snapshot": str(workbook),
                "sha256": sha256(workbook),
                "owner": "regression-test",
                "format": "xlsx",
                "data_classification": "internal",
                "derived_from": [],
            }
        ],
    }
    inventory_path = root / "source-inventory-clean.json"
    write_json(inventory_path, inventory)
    record(
        results,
        "source inventory accepts hashed authoritative inputs",
        execute(
            [
                checker,
                "--inventory",
                inventory_path,
                "--output",
                root / "source-inventory-clean-report.json",
            ]
        ),
        0,
    )
    broken = json.loads(json.dumps(inventory))
    broken["sources"].append(dict(broken["sources"][0]))
    broken_path = root / "source-inventory-duplicate-id.json"
    write_json(broken_path, broken)
    record(
        results,
        "source inventory rejects duplicate source identities",
        execute(
            [
                checker,
                "--inventory",
                broken_path,
                "--output",
                root / "source-inventory-duplicate-id-report.json",
            ]
        ),
        1,
    )


def duplicate_tests(root: Path, results: list[dict]) -> None:
    workbook_path = root / "duplicates.xlsx"
    workbook = Workbook()
    raw = workbook.active
    raw.title = "Raw Data"
    raw.append(["row_id", "amount"])
    raw.append([1, 10])
    raw.append([1, 10])
    notes = workbook.create_sheet("Notes")
    notes.append(["note"])
    notes.append(["clean"])
    workbook.save(workbook_path)
    config = {
        "check_normalized_sheet_names": True,
        "excluded_sheets": [],
        "sheets": [
            {
                "name": "Notes",
                "role": "other",
                "header_row": 1,
                "check_duplicate_headers": True,
                "check_exact_rows": True,
                "key_columns": [],
                "allowed_duplicate_keys": [],
            }
        ],
    }
    config_path = root / "duplicates-omitted-sheet.json"
    report_path = root / "duplicates-report.json"
    write_json(config_path, config)
    record(
        results,
        "duplicate audit rejects uncovered workbook sheets",
        execute(
            [
                BOOST / "check_workbook_duplicates.py",
                workbook_path,
                "--config",
                config_path,
                "--report",
                report_path,
            ]
        ),
        1,
    )
    clean_workbook_path = root / "duplicates-clean.xlsx"
    clean_workbook = Workbook()
    clean_raw = clean_workbook.active
    clean_raw.title = "Raw Data"
    clean_raw.append(["row_id", "amount"])
    clean_raw.append([1, 10])
    clean_raw.append([2, 20])
    clean_notes = clean_workbook.create_sheet("Notes")
    clean_notes.append(["note"])
    clean_notes.append(["clean"])
    clean_workbook.save(clean_workbook_path)
    clean_config = {
        "check_normalized_sheet_names": True,
        "excluded_sheets": [],
        "sheets": [
            {
                "name": "Raw Data",
                "role": "raw",
                "header_row": 1,
                "check_duplicate_headers": True,
                "check_exact_rows": True,
                "key_columns": ["row_id"],
                "ignore_columns": [],
                "allowed_duplicate_keys": [],
            },
            {
                "name": "Notes",
                "role": "other",
                "header_row": 1,
                "check_duplicate_headers": True,
                "check_exact_rows": True,
                "key_columns": [],
                "ignore_columns": [],
                "allowed_duplicate_keys": [],
            },
        ],
    }
    clean_config_path = root / "duplicates-clean.json"
    write_json(clean_config_path, clean_config)
    record(
        results,
        "duplicate audit accepts complete clean sheet coverage",
        execute(
            [
                BOOST / "check_workbook_duplicates.py",
                clean_workbook_path,
                "--config",
                clean_config_path,
                "--report",
                root / "duplicates-clean-report.json",
            ]
        ),
        0,
    )
    exception_workbook_path = root / "duplicates-reviewed-exception.xlsx"
    exception_workbook = Workbook()
    exception_raw = exception_workbook.active
    exception_raw.title = "Raw Data"
    exception_raw.append(["row_id", "amount"])
    exception_raw.append([1, 10])
    exception_raw.append([1, 20])
    exception_workbook.save(exception_workbook_path)
    exception_config = {
        "check_normalized_sheet_names": True,
        "excluded_sheets": [],
        "sheets": [
            {
                "name": "Raw Data",
                "role": "raw",
                "header_row": 1,
                "check_duplicate_headers": True,
                "check_exact_rows": True,
                "key_columns": ["row_id"],
                "ignore_columns": [],
                "allowed_duplicate_keys": [
                    {
                        "values": [1],
                        "reason": "Two reviewed source installments share one transaction ID.",
                        "owner": "regression-test",
                        "evidence": "synthetic-source-contract",
                        "checked_at": NOW,
                    }
                ],
            }
        ],
    }
    exception_config_path = root / "duplicates-reviewed-exception.json"
    write_json(exception_config_path, exception_config)
    record(
        results,
        "duplicate audit accepts only a fully evidenced key exception",
        execute(
            [
                BOOST / "check_workbook_duplicates.py",
                exception_workbook_path,
                "--config",
                exception_config_path,
                "--report",
                root / "duplicates-reviewed-exception-report.json",
            ]
        ),
        0,
    )


def foreign_tests(root: Path, results: list[dict]) -> None:
    data = root / "foreign.csv"
    write_csv(
        data,
        [
            ["row_id", "year", "funding_code", "is_foreign", "approved"],
            [1, 2023, "10", "false", 100],
            [2, 2023, "27 A", "true", 50],
        ],
    )
    predicate = root / "foreign-predicate.json"
    write_json(
        predicate,
        {
            "description": "Raw funding codes beginning with 27 are foreign.",
            "branches": [
                {
                    "all": [
                        {
                            "column": "funding_code",
                            "operator": "starts_with",
                            "value": "27 ",
                        }
                    ]
                }
            ],
        },
    )
    common = [
        BOOST / "check_foreign_funding.py",
        "--data",
        data,
        "--flag-column",
        "is_foreign",
        "--predicate-config",
        predicate,
        "--require-independent-predicate",
        "--year-column",
        "year",
        "--measure",
        "approved",
        "--id-column",
        "row_id",
    ]
    record(
        results,
        "foreign audit independently derives the expected flag",
        execute([*common, "--report", root / "foreign-report.json"]),
        0,
    )
    copied = root / "foreign-copied-predicate.json"
    write_json(
        copied,
        {
            "description": "Invalid copied-output predicate.",
            "branches": [
                {
                    "all": [
                        {
                            "column": "is_foreign",
                            "operator": "equals",
                            "value": "true",
                        }
                    ]
                }
            ],
        },
    )
    copied_command = [
        str(item) if item != predicate else str(copied) for item in common
    ]
    record(
        results,
        "foreign audit rejects predicates based on the output flag",
        execute([*copied_command, "--report", root / "foreign-copied-report.json"]),
        1,
    )


def overcounting_tests(root: Path, results: list[dict]) -> None:
    raw = root / "raw.csv"
    rules = root / "rules.csv"
    dictionary = root / "dictionary.csv"
    config = root / "overlap-config.json"
    write_csv(raw, [["year", "approved", "econ1"], [2023, 100, "A"]])
    write_csv(
        rules,
        [["code", "row_type", "sheet", "measure", "criteria_json"]],
    )
    write_csv(
        dictionary,
        [
            ["code", "tag_kind", "econ", "econ_sub", "func", "func_sub"],
            ["A", "primary", "Economic affairs", "", "", ""],
        ],
    )
    write_json(
        config,
        {
            "country": "Auditland",
            "expected_stages": ["Approved"],
            "minimum_rules_per_stage": 1,
            "expected_codes_by_stage": {"Approved": ["A"]},
            "ranges": [
                {
                    "name": "all-years",
                    "input": str(raw),
                    "format": "csv",
                    "measures": ["approved"],
                    "measure_dispatch": ["approved"],
                    "field_map": {"year": "year", "econ1": "econ1"},
                }
            ],
        },
    )
    record(
        results,
        "overcounting audit rejects empty rule coverage",
        execute(
            [
                OVERCOUNTING / "detect_tag_overlaps.py",
                "--config",
                config,
                "--tag-rules",
                rules,
                "--code-dictionary",
                dictionary,
                "--output",
                root / "overlap-report.json",
            ]
        ),
        1,
    )
    rule_header = [
        "code",
        "row_type",
        "sheet",
        "measure",
        "criteria_json",
        "sample_formula",
        "formula_cell",
    ]
    criterion = {"field": "econ1", "op": "=", "value": "A"}
    write_csv(
        rules,
        [
            rule_header,
            [
                "A",
                "TAG",
                "Approved",
                "approved",
                json.dumps([criterion]),
                "=SUMIFS(...) ",
                "B2",
            ],
        ],
    )
    clean_command = [
        OVERCOUNTING / "detect_tag_overlaps.py",
        "--config",
        config,
        "--tag-rules",
        rules,
        "--code-dictionary",
        dictionary,
        "--output",
        root / "overlap-clean-report.json",
    ]
    record(
        results,
        "overcounting audit accepts complete non-overlapping coverage",
        execute(clean_command),
        0,
    )
    write_csv(
        rules,
        [
            rule_header,
            [
                "A",
                "TAG",
                "Approved",
                "approved",
                json.dumps([[criterion], [criterion]]),
                "=SUMIFS(...)+SUMIFS(...)",
                "B2",
            ],
        ],
    )
    record(
        results,
        "overcounting audit rejects within-formula self-double",
        execute(
            [
                *clean_command[:-1],
                root / "overlap-self-double-report.json",
            ]
        ),
        1,
    )


def subnational_tests(root: Path, results: list[dict]) -> None:
    boundaries = root / "boundaries.geojson"
    write_json(
        boundaries,
        {
            "type": "FeatureCollection",
            "features": [
                {
                    "type": "Feature",
                    "properties": {"admin1_region": "Only Region"},
                    "geometry": {"type": "Point", "coordinates": [0, 0]},
                }
            ],
        },
    )
    contract = root / "admin-empty.json"
    write_json(
        contract,
        {
            "country": "Auditland",
            "iso3": "AUD",
            "subnational_required": True,
            "required_dataset_names": ["population", "boost_geography"],
            "target": {
                "admin_level": "admin1",
                "vintage": "2026",
                "boundary_path": str(boundaries),
                "boundary_format": "geojson",
                "boundary_name_property": "admin1_region",
                "expected_unit_count": 1,
                "require_geometry_validation": True,
            },
            "datasets": [],
        },
    )
    record(
        results,
        "subnational audit rejects missing required datasets",
        execute(
            [
                SUBNATIONAL / "audit_admin_coverage.py",
                "--contract",
                contract,
                "--output",
                root / "admin-report.json",
            ]
        ),
        1,
    )

    reviewed_boundaries = root / "boundaries-reviewed-no-data.geojson"
    write_json(
        reviewed_boundaries,
        {
            "type": "FeatureCollection",
            "features": [
                {
                    "type": "Feature",
                    "properties": {"admin1_region": "North"},
                    "geometry": {"type": "Point", "coordinates": [0, 0]},
                },
                {
                    "type": "Feature",
                    "properties": {"admin1_region": "Central"},
                    "geometry": {"type": "Point", "coordinates": [1, 0]},
                },
            ],
        },
    )
    population = root / "population-no-data.csv"
    write_csv(population, [["region", "year", "population"], ["North", 2023, 10]])
    base_contract = {
        "country": "Auditland",
        "iso3": "AUD",
        "subnational_required": True,
        "required_dataset_names": ["population"],
        "target": {
            "admin_level": "admin1",
            "vintage": "2026",
            "boundary_path": str(reviewed_boundaries),
            "boundary_format": "geojson",
            "boundary_name_property": "admin1_region",
            "expected_unit_count": 2,
            "require_geometry_validation": True,
        },
    }
    unreviewed = dict(base_contract)
    unreviewed["datasets"] = [
        {
            "name": "population",
            "role": "population",
            "path": str(population),
            "format": "csv",
            "name_column": "region",
            "year_column": "year",
            "value_column": "population",
            "required_complete": True,
            "required_years": [2023],
            "accepted_no_data": [
                {"target": "Central", "reason": "No source row observed."}
            ],
            "mapping": {},
        }
    ]
    unreviewed_path = root / "admin-unreviewed-no-data.json"
    unreviewed_report = root / "admin-unreviewed-no-data-report.json"
    write_json(unreviewed_path, unreviewed)
    unreviewed_run = execute(
        [
            SUBNATIONAL / "audit_admin_coverage.py",
            "--contract",
            unreviewed_path,
            "--output",
            unreviewed_report,
        ]
    )
    unreviewed_result = json.loads(unreviewed_report.read_text(encoding="utf-8"))
    unreviewed_issues = unreviewed_result["datasets"][0]["issues"]
    record_condition(
        results,
        "subnational audit rejects an unreviewed no-data exception",
        unreviewed_result["datasets"][0]["passed"] is False
        and any(".owner is required" in issue for issue in unreviewed_issues)
        and any(".evidence is required" in issue for issue in unreviewed_issues)
        and any(".checked_at" in issue for issue in unreviewed_issues),
        unreviewed_run,
    )

    reviewed = json.loads(json.dumps(base_contract))
    reviewed["datasets"] = json.loads(json.dumps(unreviewed["datasets"]))
    reviewed["datasets"][0]["accepted_no_data"][0].update(
        {
            "owner": "geography-reviewer",
            "evidence": "official-source-coverage-table",
            "checked_at": NOW,
        }
    )
    reviewed_path = root / "admin-reviewed-no-data.json"
    reviewed_report = root / "admin-reviewed-no-data-report.json"
    write_json(reviewed_path, reviewed)
    reviewed_run = execute(
        [
            SUBNATIONAL / "audit_admin_coverage.py",
            "--contract",
            reviewed_path,
            "--output",
            reviewed_report,
        ]
    )
    reviewed_result = json.loads(reviewed_report.read_text(encoding="utf-8"))
    reviewed_dataset = reviewed_result["datasets"][0]
    record_condition(
        results,
        "subnational audit accepts a fully reviewed no-data exception",
        reviewed_dataset["passed"] is True
        and reviewed_dataset["accepted_no_data"][0]["owner"] == "geography-reviewer",
        reviewed_run,
    )


def reconciliation_tests(root: Path, results: list[dict]) -> None:
    left = root / "left.csv"
    right = root / "right.csv"
    write_csv(left, [["year", "amount"], [2023, 40], [2023, 60]])
    write_csv(right, [["year", "amount"], [2023, 100]])
    record(
        results,
        "reconciliation rejects duplicate keys before aggregation",
        execute(
            [
                VALIDATION / "reconcile_csv.py",
                "--left",
                left,
                "--right",
                right,
                "--keys",
                "year",
                "--left-value",
                "amount",
                "--right-value",
                "amount",
                "--detail-csv",
                root / "reconcile-detail.csv",
                "--report",
                root / "reconcile-report.json",
            ]
        ),
        1,
    )
    clean_left = root / "left-clean.csv"
    write_csv(clean_left, [["year", "amount"], [2023, 100]])
    record(
        results,
        "reconciliation accepts unique matching inputs",
        execute(
            [
                VALIDATION / "reconcile_csv.py",
                "--left",
                clean_left,
                "--right",
                right,
                "--keys",
                "year",
                "--left-value",
                "amount",
                "--right-value",
                "amount",
                "--detail-csv",
                root / "reconcile-clean-detail.csv",
                "--report",
                root / "reconcile-clean-report.json",
            ]
        ),
        0,
    )


def demo_tests(root: Path, results: list[dict]) -> None:
    demo_output = root / "demoland"
    command = [
        SUITE_ROOT / "mega-country-onboarding" / "scripts" / "run_demoland.py",
        "--output",
        demo_output,
    ]
    record(
        results,
        "DemoLand creates a valid first-hour onboarding workspace",
        execute(command),
        0,
    )


def main() -> int:
    results: list[dict] = []
    with tempfile.TemporaryDirectory(prefix="mega-onboarding-regression-") as temp:
        root = Path(temp)
        source_inventory_tests(root, results)
        manifest_tests(root, results)
        duplicate_tests(root, results)
        foreign_tests(root, results)
        overcounting_tests(root, results)
        subnational_tests(root, results)
        reconciliation_tests(root, results)
        demo_tests(root, results)
    report = {
        "passed": all(item["passed"] for item in results),
        "tests": len(results),
        "failed": [item for item in results if not item["passed"]],
        "results": results,
    }
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0 if report["passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
